#!/usr/bin/env python3
"""Generate ET6 startup config workbooks from a role-based deployment spec.

Usage:
  python generate_startup_config.py --template-dir <dir> --output-dir <dir> --spec <spec.json>

The spec is intentionally role-based so callers do not need to hand-author every row:
{
  "machines": [
    {"id": 10, "role": "routing", "inner_ip": "10.0.0.10", "public_ip": "60.204.147.173"},
    {"id": 20, "role": "zone", "inner_ip": "10.0.0.20"},
    {"id": 30, "role": "load_balance", "inner_ip": "10.0.0.30"},
    {"id": 99, "role": "global", "inner_ip": "10.0.0.90"}
  ],
  "zones": [100, 101],
  "routing_node_processes_per_machine": 2,
  "node_counts": {"LogicNodeServer": 2, "GateNodeServer": 2, "BattleNodeServer": 2, "MapNodeServer": 2},
  "db": {"connection": "10.0.0.90:27017", "account": "admin", "password": "admin"}
}
"""

from __future__ import annotations

import argparse
import copy
import json
import shutil
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


FILES = [
    "StartMachineConfig@s.xlsx",
    "StartProcessConfig@s.xlsx",
    "StartSceneConfig@s.xlsx",
    "StartZoneConfig@s.xlsx",
    "StartBalanceConfig@s.xlsx",
    "StartMergeServerConfig@s.xlsx",
]

GLOBAL_MANAGER_TYPES = {"RealmBalanceServer", "BattleBalanceServer", "MapNodeBalanceServer"}
NODE_TYPES = {"LogicNodeServer", "GateNodeServer", "BattleNodeServer", "MapNodeServer"}
ZONE_PROCESS_MAP = {"LocalUserServer": 0, "MapServer": 1, "UnitCache": 3}


def as_int(value: Any) -> int | None:
    if value in (None, ""):
        return None
    return int(value)


def key_to_col(ws) -> dict[str, int]:
    return {
        ws.cell(row=4, column=col).value: col
        for col in range(1, ws.max_column + 1)
        if ws.cell(row=4, column=col).value
    }


def read_rows(ws) -> list[dict[str, Any]]:
    cols = key_to_col(ws)
    rows: list[dict[str, Any]] = []
    for row_idx in range(6, ws.max_row + 1):
        item: dict[str, Any] = {}
        for key, col in cols.items():
            value = ws.cell(row=row_idx, column=col).value
            if value not in (None, ""):
                item[key] = value
        if item:
            rows.append(item)
    return rows


def copy_row_style(ws, source_row: int, target_row: int) -> None:
    for col in range(1, ws.max_column + 1):
        src = ws.cell(row=source_row, column=col)
        dst = ws.cell(row=target_row, column=col)
        if src.has_style:
            dst._style = copy.copy(src._style)
        dst.number_format = src.number_format
        dst.font = copy.copy(src.font)
        dst.fill = copy.copy(src.fill)
        dst.border = copy.copy(src.border)
        dst.alignment = copy.copy(src.alignment)
        dst.protection = copy.copy(src.protection)
    if ws.row_dimensions[source_row].height is not None:
        ws.row_dimensions[target_row].height = ws.row_dimensions[source_row].height


def clear_data(ws, min_rows: int) -> None:
    target_max = max(ws.max_row, 5 + min_rows)
    for row in range(6, target_max + 1):
        if row > ws.max_row:
            copy_row_style(ws, 6, row)
        for col in range(1, ws.max_column + 1):
            ws.cell(row=row, column=col).value = None


def write_rows(ws, rows: list[dict[str, Any]]) -> None:
    clear_data(ws, len(rows))
    cols = key_to_col(ws)
    for row_idx, row_data in enumerate(rows, start=6):
        if row_idx != 6:
            copy_row_style(ws, 6, row_idx)
        for key, value in row_data.items():
            if key in cols:
                ws.cell(row=row_idx, column=cols[key]).value = value


def require_one(items: list[dict[str, Any]], role: str) -> dict[str, Any]:
    matches = [item for item in items if item["role"] == role]
    if len(matches) != 1:
        raise ValueError(f"expected exactly one {role!r} machine, got {len(matches)}")
    return matches[0]


def normalize_spec(spec: dict[str, Any]) -> dict[str, Any]:
    machines = spec.get("machines") or []
    if not machines:
        raise ValueError("spec.machines is required")
    for machine in machines:
        machine["id"] = int(machine["id"])
        machine["role"] = str(machine["role"]).lower()
        if "inner_ip" not in machine:
            raise ValueError(f"machine {machine['id']} missing inner_ip")
    spec["machines"] = machines
    spec["zones"] = [int(zone) for zone in (spec.get("zones") or [])]
    if not spec["zones"]:
        raise ValueError("spec.zones is required")
    spec.setdefault("routing_node_processes_per_machine", 2)
    spec.setdefault(
        "node_counts",
        {"LogicNodeServer": 2, "GateNodeServer": 2, "BattleNodeServer": 2, "MapNodeServer": 2},
    )
    spec.setdefault("db", {})
    spec["db"].setdefault("account", "admin")
    spec["db"].setdefault("password", "admin")
    global_machine = require_one(machines, "global")
    spec["db"].setdefault("connection", f"{global_machine['inner_ip']}:27017")
    return spec


def sheet_by_range(wb, marker: str):
    for ws in wb.worksheets:
        if marker in ws.title:
            return ws
    raise KeyError(f"sheet containing {marker!r} not found")


def copy_templates(template_dir: Path, output_dir: Path) -> None:
    output_dir.mkdir(parents=True, exist_ok=True)
    for name in FILES:
        src = template_dir / name
        if not src.exists():
            raise FileNotFoundError(src)
        shutil.copy2(src, output_dir / name)


def load_scene_templates(template_dir: Path) -> dict[str, list[dict[str, Any]]]:
    wb = load_workbook(template_dir / "StartSceneConfig@s.xlsx", data_only=False)
    return {ws.title: read_rows(ws) for ws in wb.worksheets}


def scene_row(template: dict[str, Any], **updates: Any) -> dict[str, Any]:
    row = dict(template)
    row.update(updates)
    return row


def process_row(process_id: int, machine_id: int) -> dict[str, int]:
    return {"Id": process_id, "MachineId": machine_id, "InnerPort": process_id + 10000}


def generate(template_dir: Path, output_dir: Path, spec: dict[str, Any]) -> dict[str, Any]:
    spec = normalize_spec(spec)
    copy_templates(template_dir, output_dir)
    scene_templates = load_scene_templates(template_dir)

    global_machine = require_one(spec["machines"], "global")
    zone_machine = require_one(spec["machines"], "zone")
    lb_machine = require_one(spec["machines"], "load_balance")
    routing_machines = [m for m in spec["machines"] if m["role"] == "routing"]
    if not routing_machines:
        raise ValueError("at least one routing machine is required")

    processes: dict[int, dict[str, Any]] = {}
    scene_sheets: dict[str, list[dict[str, Any]]] = defaultdict(list)

    def add_scene(sheet: str, row: dict[str, Any], machine_id: int) -> None:
        scene_sheets[sheet].append(row)
        process_id = as_int(row.get("Process"))
        if process_id is not None:
            processes[process_id] = process_row(process_id, machine_id)

    global_system_sheet = next(name for name in scene_templates if "10000-10249" in name)
    global_business_sheet = next(name for name in scene_templates if "10250-10499" in name)
    control_sheet = next(name for name in scene_templates if "12500-19999" in name)
    routing_sheet = next(name for name in scene_templates if "10500-10999" in name)
    logic_sheet = next(name for name in scene_templates if "11000-11499" in name)
    gate_sheet = next(name for name in scene_templates if "11500-11999" in name)
    battle_sheet = next(name for name in scene_templates if "12000-12249" in name)
    map_sheet = next(name for name in scene_templates if "12300-12399" in name)
    game_zone_sheet = next(name for name in scene_templates if "30000+" in name)

    for template in scene_templates[global_system_sheet]:
        scene_id = as_int(template.get("Id"))
        if scene_id is None:
            continue
        machine_id = global_machine["id"]
        if template.get("SceneType") == "RouterManager":
            machine_id = routing_machines[0]["id"]
        add_scene(global_system_sheet, template, machine_id)

    for template in scene_templates[global_business_sheet]:
        scene_id = as_int(template.get("Id"))
        if scene_id is None:
            continue
        add_scene(global_business_sheet, template, global_machine["id"])

    routing_template = scene_templates[routing_sheet][0]
    routing_id = 10500
    for machine in routing_machines:
        for _ in range(int(spec["routing_node_processes_per_machine"])):
            add_scene(
                routing_sheet,
                scene_row(
                    routing_template,
                    Id=routing_id,
                    Process=routing_id,
                    Zone=2,
                    SceneType="RoutingNodeServer",
                    SceneDesc=f"Routing node server {routing_id}",
                    Name=f"RoutingNodeServer{routing_id}",
                    OuterPort=routing_id,
                ),
                machine["id"],
            )
            routing_id += 1

    node_sheet_info = {
        "LogicNodeServer": (logic_sheet, 11000, ""),
        "GateNodeServer": (gate_sheet, 11500, None),
        "BattleNodeServer": (battle_sheet, 12000, ""),
        "MapNodeServer": (map_sheet, 12300, ""),
    }
    for scene_type, (sheet, start_id, outer) in node_sheet_info.items():
        template = scene_templates[sheet][0]
        for index in range(int(spec["node_counts"].get(scene_type, 0))):
            scene_id = start_id + index
            add_scene(
                sheet,
                scene_row(
                    template,
                    Id=scene_id,
                    Process=scene_id,
                    Zone=2,
                    SceneType=scene_type,
                    SceneDesc=f"{scene_type} {scene_id}",
                    Name=f"{scene_type}{scene_id}",
                    OuterPort=(scene_id if outer is None else outer),
                ),
                lb_machine["id"],
            )

    control_template = scene_templates[control_sheet][0]
    for machine in sorted(spec["machines"], key=lambda item: item["id"]):
        scene_id = 12500 + machine["id"]
        add_scene(
            control_sheet,
            scene_row(
                control_template,
                Id=scene_id,
                Process=scene_id,
                Zone=3,
                SceneType="MachineControlServer",
                LinkMachineId=machine["id"],
                SceneDesc=f"Machine control server for machine {machine['id']}",
                Name=f"MachineControlServer{scene_id}",
                OuterPort="",
            ),
            machine["id"],
        )

    game_templates = scene_templates[game_zone_sheet]
    for zone_index, zone in enumerate(spec["zones"], start=1):
        zone_base_scene_id = 30000 + (zone_index - 1) * 100
        for template in game_templates:
            template_id = as_int(template.get("Id"))
            if template_id is None:
                continue
            scene_type = template.get("SceneType")
            scene_id = zone_base_scene_id + (template_id - 30000)
            process_id = zone_base_scene_id + ZONE_PROCESS_MAP.get(scene_type, 2)
            add_scene(
                game_zone_sheet,
                scene_row(
                    template,
                    Id=scene_id,
                    Process=process_id,
                    Zone=zone,
                    SceneType=scene_type,
                    GameZoneDisplayName=f"Zone{zone}",
                    SceneDesc=f"{scene_type} for zone {zone}",
                    Name=f"{scene_type}-{zone_index}",
                    OuterPort=(scene_id if scene_type == "LocalUserServer" else 0),
                    InnerZoneIndex=zone_index,
                    BackendZoneId=1000000 + zone,
                    ChannelId=1000,
                    ServerInnerName=f"Zone{zone}",
                    LinkMachineId=zone_machine["id"],
                ),
                zone_machine["id"],
            )

    machine_rows = []
    for machine in sorted(spec["machines"], key=lambda item: item["id"]):
        inner_ip = machine["inner_ip"]
        out_real_ip = machine.get("public_ip") or inner_ip
        machine_rows.append(
            {
                "Id": machine["id"],
                "InnerIP": inner_ip,
                "OuterIP": inner_ip,
                "OutRealIP": out_real_ip,
                "EnableGameZone": 1 if machine["role"] == "zone" else 0,
            }
        )

    wb = load_workbook(output_dir / "StartMachineConfig@s.xlsx")
    write_rows(wb["StartMachineConfig"], machine_rows)
    wb.save(output_dir / "StartMachineConfig@s.xlsx")

    process_rows = [processes[pid] for pid in sorted(processes)]
    wb = load_workbook(output_dir / "StartProcessConfig@s.xlsx")
    write_rows(wb["StartProcessConfig"], process_rows)
    wb.save(output_dir / "StartProcessConfig@s.xlsx")

    wb = load_workbook(output_dir / "StartSceneConfig@s.xlsx")
    for ws in wb.worksheets:
        write_rows(ws, scene_sheets.get(ws.title, []))
    wb.save(output_dir / "StartSceneConfig@s.xlsx")

    zone_rows = [
        {"Id": 1, "DBConnection": spec["db"]["connection"], "DBName": "AccountDatabase", "Account": spec["db"]["account"], "Password": spec["db"]["password"], "Desc": "Account database"},
        {"Id": 2, "DBConnection": spec["db"]["connection"], "DBName": "GlobalDatabase", "Account": spec["db"]["account"], "Password": spec["db"]["password"], "Desc": "Global database"},
        {"Id": 3, "DBConnection": spec["db"]["connection"], "DBName": "ControlDatabase", "Account": spec["db"]["account"], "Password": spec["db"]["password"], "Desc": "Control database"},
    ]
    for zone in spec["zones"]:
        zone_rows.append({"Id": zone, "DBConnection": spec["db"]["connection"], "DBName": f"GameZone{zone}", "Account": spec["db"]["account"], "Password": spec["db"]["password"], "Desc": f"Game zone {zone}"})
    wb = load_workbook(output_dir / "StartZoneConfig@s.xlsx")
    write_rows(wb["StartZoneConfig"], zone_rows)
    wb.save(output_dir / "StartZoneConfig@s.xlsx")

    balance_rows = [
        {"Id": 1, "SceneType": "LogicNodeServer", "Strategy": "LeastActive"},
        {"Id": 2, "SceneType": "GateNodeServer", "Strategy": "LeastActive"},
        {"Id": 3, "SceneType": "BattleNodeServer", "Strategy": "LeastActive"},
        {"Id": 4, "SceneType": "MapNodeServer", "Strategy": "LeastActive"},
    ]
    wb = load_workbook(output_dir / "StartBalanceConfig@s.xlsx")
    write_rows(wb.worksheets[0], balance_rows)
    wb.save(output_dir / "StartBalanceConfig@s.xlsx")

    wb = load_workbook(output_dir / "StartMergeServerConfig@s.xlsx")
    write_rows(wb.worksheets[0], [])
    wb.save(output_dir / "StartMergeServerConfig@s.xlsx")

    validation = validate(output_dir, global_machine["id"])
    return {
        "output_dir": str(output_dir),
        "machines": len(machine_rows),
        "processes": len(process_rows),
        "scenes": sum(len(rows) for rows in scene_sheets.values()),
        "zones": len(zone_rows),
        "warnings": validation,
    }


def workbook_rows(path: Path, sheet_name: str | None = None) -> dict[str, list[dict[str, Any]]]:
    wb = load_workbook(path, data_only=False)
    sheets = [wb[sheet_name]] if sheet_name else wb.worksheets
    return {ws.title: read_rows(ws) for ws in sheets}


def validate(output_dir: Path, global_machine_id: int) -> list[str]:
    warnings: list[str] = []
    machines = workbook_rows(output_dir / "StartMachineConfig@s.xlsx", "StartMachineConfig")["StartMachineConfig"]
    processes = workbook_rows(output_dir / "StartProcessConfig@s.xlsx", "StartProcessConfig")["StartProcessConfig"]
    zones = workbook_rows(output_dir / "StartZoneConfig@s.xlsx", "StartZoneConfig")["StartZoneConfig"]
    scenes_by_sheet = workbook_rows(output_dir / "StartSceneConfig@s.xlsx")
    scenes = [row for rows in scenes_by_sheet.values() for row in rows]

    machine_ids = {as_int(row.get("Id")) for row in machines}
    process_ids = {as_int(row.get("Id")) for row in processes}
    zone_ids = {as_int(row.get("Id")) for row in zones}
    process_machine = {as_int(row.get("Id")): as_int(row.get("MachineId")) for row in processes}

    for label, data in [("machine", machines), ("process", processes), ("zone", zones), ("scene", scenes)]:
        ids = [as_int(row.get("Id")) for row in data if as_int(row.get("Id")) is not None]
        dupes = sorted(item for item, count in Counter(ids).items() if count > 1)
        if dupes:
            warnings.append(f"duplicate {label} ids: {dupes}")

    for row in machines:
        if row.get("OuterIP") != row.get("InnerIP"):
            warnings.append(f"machine {row.get('Id')} OuterIP does not match InnerIP")

    zone_enabled = [as_int(row.get("Id")) for row in machines if as_int(row.get("EnableGameZone")) == 1]
    if len(zone_enabled) != 1:
        warnings.append(f"expected exactly one EnableGameZone machine, got {zone_enabled}")

    missing_machines = sorted({as_int(row.get("MachineId")) for row in processes if as_int(row.get("MachineId")) not in machine_ids})
    if missing_machines:
        warnings.append(f"missing machine references: {missing_machines}")

    missing_processes = sorted({as_int(row.get("Process")) for row in scenes if as_int(row.get("Process")) is not None and as_int(row.get("Process")) not in process_ids})
    if missing_processes:
        warnings.append(f"missing process references: {missing_processes}")

    missing_zones = sorted({as_int(row.get("Zone")) for row in scenes if as_int(row.get("Zone")) is not None and as_int(row.get("Zone")) not in zone_ids})
    if missing_zones:
        warnings.append(f"missing zone references: {missing_zones}")

    for row in scenes:
        if row.get("SceneType") in GLOBAL_MANAGER_TYPES and process_machine.get(as_int(row.get("Process"))) != global_machine_id:
            warnings.append(f"{row.get('SceneType')} process {row.get('Process')} is not on global machine {global_machine_id}")

    inner_ports: dict[tuple[int, int], list[int]] = defaultdict(list)
    for row in processes:
        machine_id = as_int(row.get("MachineId"))
        inner_port = as_int(row.get("InnerPort"))
        if machine_id is not None and inner_port is not None:
            inner_ports[(machine_id, inner_port)].append(as_int(row.get("Id")))
    for key, values in inner_ports.items():
        if len(values) > 1:
            warnings.append(f"inner port collision {key}: {values}")

    outer_ports: dict[int, list[int]] = defaultdict(list)
    for row in scenes:
        outer_port = as_int(row.get("OuterPort"))
        if outer_port and outer_port > 0:
            outer_ports[outer_port].append(as_int(row.get("Id")))
    for port, values in outer_ports.items():
        if len(values) > 1:
            warnings.append(f"outer port collision {port}: {values}")
    return warnings


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--template-dir", required=True, type=Path)
    parser.add_argument("--output-dir", required=True, type=Path)
    parser.add_argument("--spec", required=True, type=Path)
    args = parser.parse_args()
    spec = json.loads(args.spec.read_text(encoding="utf-8"))
    summary = generate(args.template_dir, args.output_dir, spec)
    print(json.dumps(summary, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
